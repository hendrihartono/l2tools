import pandas as pd
import json
from datetime import datetime, timedelta
import os
import openpyxl

class ProdIssue:
    def __init__(self):
        self.response_body_column = 'response.body'
        self.timestamp_column = '@timestamp'
        self.output_file = f'{(datetime.now() - timedelta(days=1)).strftime("%m%d%Y")}.xlsx'
        self.file_path = 'temporary-file.xlsx'
        self.detail_error_column = 'Error Detail by Confluence'
        self.error_mapping_code_sheet = 'mapping-error-code'
        self.target_file_path = self.output_file
        self.source_file_path = f'{self.error_mapping_code_sheet}.xlsx'
        self.new_sheet_name = self.error_mapping_code_sheet

    def creating_output(self):
        input_file = input("Masukkan nama file download gan tanpa .csv (or type 'back' to return to menu): ")
        if input_file.lower() == 'back':
            return 'back'
        file_path = f'./{input_file}.csv'
        
        try:
            df = pd.read_csv(file_path, delimiter='|')
        except FileNotFoundError:
            print(f"Yang bener nama filenya, Mang... -____-")
            return  

        df[self.response_body_column] = df[self.response_body_column].apply(self.parse_json)

        df = df[df[self.response_body_column].apply(lambda x: isinstance(x, dict))]

        df['errorCode'] = df[self.response_body_column].apply(lambda x: x.get('errorCode'))
        df['errorDesc'] = df[self.response_body_column].apply(lambda x: x.get('errorDesc'))

        grouped_df = df.groupby(['app_name', 'errorCode', 'errorDesc', 'request.uri']).size().reset_index(name='count')

        app_groups = grouped_df.groupby('app_name')

        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            for app_name, group in app_groups:
                sheet_name = app_name[:31]  
                group.drop(columns='app_name', inplace=True) 
                group.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"Mang Cecep sedang memasak {sheet_name}....")
        
    def parse_json(self, json_str):
        if isinstance(json_str, str):
            try:
                return json.loads(json_str.replace('""', '"'))
            except json.JSONDecodeError:
                return None
        return None

    def add_sheet_from_file(self):
        try:
            target_wb = openpyxl.load_workbook(self.target_file_path)
        except FileNotFoundError:
            print(f"Error: The target file {self.target_file_path} was not found.")
            return
        except Exception as e:
            print(f"Error loading target workbook: {e}")
            return
        try:
            source_wb = openpyxl.load_workbook(self.source_file_path)
            source_ws = source_wb.active
        except FileNotFoundError:
            print(f"Error: The source file {self.source_file_path} was not found.")
            return
        except Exception as e:
            print(f"Error loading source workbook: {e}")
            return

        if self.new_sheet_name in target_wb.sheetnames:
            print(f"Error: The sheet '{self.new_sheet_name}' already exists in the target file.")
            return

        target_ws = target_wb.create_sheet(title=self.new_sheet_name)

        for row_index, row in enumerate(source_ws.iter_rows(values_only=True), start=1):
            for col_index, cell_value in enumerate(row, start=1):
                target_ws.cell(row=row_index, column=col_index, value=cell_value)

        target_wb.save(self.file_path)
        
    def add_vlookup_column_to_all_sheets(self):
        try:
            wb = openpyxl.load_workbook(self.file_path)
        except FileNotFoundError:
            print(f"Error: The file {self.file_path} was not found.")
            return
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            last_column = ws.max_column
            new_column_index = last_column + 1

            ws.cell(row=1, column=new_column_index, value=self.detail_error_column)

            for row in range(2, ws.max_row + 1):
                formula = f'=VLOOKUP(A{row}, \'{self.error_mapping_code_sheet}\'!A:B, 2, FALSE)'
                ws.cell(row=row, column=new_column_index, value=formula)

        wb.save(self.output_file)
        print(f'Mang Cecep selesai memasak {self.output_file}! :D')    
        try:
            os.remove(self.file_path)
        except Exception as e:
            print(f"Warning: Could not remove temporary file: {e}")

    def run(self):
        if self.creating_output() == 'back':
            return 'back'
        self.add_sheet_from_file()
        self.add_vlookup_column_to_all_sheets()


if __name__ == "__main__":
    prod_issue = ProdIssue()
    if prod_issue.run() == 'back':
        print("Returning to tool selection...")
